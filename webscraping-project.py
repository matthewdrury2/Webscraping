from unicodedata import name
from urllib.request import urlopen, Request
from bs4 import BeautifulSoup
from twilio.rest import Client
import openpyxl as xl
from openpyxl.styles import Font

url = 'https://cryptoslate.com/coins/'
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3'}

req = Request(url, headers=headers)

webpage = urlopen(req).read()

soup = BeautifulSoup(webpage, "html.parser")

print(soup.title.text)

title = soup.title

print(title.text)
print('Ranked by Market Capitalization')
print()
####


wb = xl.Workbook()

ws = wb.active

ws.title = 'Webscraping Project'

ws['A1'] = 'Name and Symbol'
ws['B1'] = 'Current Price'
ws['C1'] = '% change in price in past 24 hours'
ws['D1'] = 'Price 24 hours ago'
ws['E1'] = 'Price change in 24 hours'


####
coin_row = soup.findAll('tr')
'''
accountSID = ''

authToken = ''

Client = Client(accountSID, authToken)

TwilioNumber = ""

mycellphone = ""
'''
for row in coin_row[1:6]:
    td = row.findAll('td')

    name = td[1].text
    print('Name and Symbol:', name)

    price = td[2].text
    print('Current Price:', price)
    price = float(td[4].text.replace(
        ',', '').replace('$', '').replace('+', '').replace('%', ''))

    change = td[3].text
    print('Percent Change in Price over the last 24 hours:', change)
    change = float(td[5].text.replace(
        '', '').replace('+', '').replace('%', ''))
    change = 1 + (change/100)
    # print(change)

    dchange = price/change
    # print(dchange)

    mon = float(price - dchange)

    poneg = 'increase'
    if mon > 0:
        poneg = 'increase'
    else:
        poneg = 'decrease'

    print('The Price 24 hours ago was:', "${:,.4f}". format(dchange))

    mon = "${:,.4f}". format(mon)

    print('Price Amount Change in last 24 hours:', mon, poneg)

    print('')

    ws['A' + str(row+1)] = name
    ws['B' + str(row+1)] = price
    ws['C' + str(row+1)] = change
    ws['D' + str(row+1)] = dchange
    ws['E' + str(row+1)] = mon

'''
    if str(td[2].text) == 'Bitcoin BTC ':
        if price < 40000:
            textmessage = Client.messages.create(
                to=mycellphone, from_=TwilioNumber, body="BTC price is below $40,000!")

            # print(textmessage.status)

    if str(td[2].text) == 'Ethereum ETH ':
        if price < 3000:
            textmessage = Client.messages.create(
                to=mycellphone, from_=TwilioNumber, body="ETH price is below $3,000!")
'''
