from unicodedata import name
from urllib.request import urlopen, Request
from bs4 import BeautifulSoup
from twilio.rest import Client
import openpyxl as xl
from openpyxl.styles import Font

url = 'https://cryptoslate.com/coins/'
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3'}

req = Request(url, headers=headers)

webpage = urlopen(req).read()

soup = BeautifulSoup(webpage, "html.parser")

print(soup.title.text)

title = soup.title

print(title.text)
print('Ranked by Market Capitalization')
print()

wb = xl.Workbook()

ws = wb.active

ws.title = 'Webscraping Project'

ws['A1'] = 'Name and Symbol'
ws['B1'] = 'Current Price'
ws['C1'] = '% change in price in past 24 hours'
ws['D1'] = 'Price 24 hours ago'
ws['E1'] = 'Price change in 24 hours'

coin_row = soup.findAll('tr')

for row_num, row in enumerate(coin_row[1:6], start=2):
    td = row.findAll('td')

    name = td[1].text
    print('Name and Symbol:', name)

    price = td[2].text
    print('Current Price:', price)

    percent_change = float(td[3].text.strip('%'))
    print('Percent Change in Price over the last 24 hours:', percent_change)

    price_24h_ago = float(td[4].text.replace(',', '').replace(
        '$', '').replace('+', '').replace('%', ''))
    print('The Price 24 hours ago was:', price_24h_ago)

    price_change = float(td[5].text.replace(',', '').replace(
        '$', '').replace('+', '').replace('%', ''))
    price_change_str = "${:,.4f}".format(price_change)
    print('Price Amount Change in last 24 hours:', price_change_str)

    ws.cell(row=row_num, column=1, value=name)
    ws.cell(row=row_num, column=2, value=price)
    ws.cell(row=row_num, column=3, value=percent_change)
    ws.cell(row=row_num, column=4, value=price_24h_ago)
    ws.cell(row=row_num, column=5, value=price_change_str)

    if name == 'Bitcoin BTC':
        if abs(price_24h_ago - float(price.replace(',', '').replace('$', ''))) >= 5:
            client = Client()
            message = client.messages \
                            .create(
                                body=f"Bitcoin price changed by more than $5. Current price is {price}",
                                from_='YOUR_TWILIO_NUMBER',
                                to='YOUR_PHONE_NUMBER'
                            )
            print(message.sid)

    if name == 'Ethereum ETH':
        if abs(price_24h_ago - float(price.replace(',', '').replace('$', ''))) >= 5:
            client = Client()
            message = client.messages \
                            .create(
                                body=f"Ethereum price changed by more than $5. Current price is {price}",
                                from_='YOUR_TWILIO_NUMBER',
                                to='YOUR_PHONE_NUMBER'
                            )
            print(message.sid)

# Apply formatting to the headers
font_header = Font(bold=True)
# for cell in ws['A1:E1']:
# cell.font = font_header

# Set column widths
ws.column_dimensions['A']
