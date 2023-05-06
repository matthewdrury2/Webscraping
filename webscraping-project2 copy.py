import requests
from bs4 import BeautifulSoup
import openpyxl
from twilio.rest import Client

# Scrape CoinMarketCap website for cryptocurrency data
url = 'https://coinmarketcap.com/'
response = requests.get(url)
soup = BeautifulSoup(response.content, 'html.parser')
cryptos = soup.find_all('tr', {'class': 'cmc-table-row'})

# Create Excel spreadsheet and write data to it
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Crypto'
ws['A1'] = 'Name'
ws['B1'] = 'Symbol'
ws['C1'] = 'Price'
ws['D1'] = '% Change (24h)'
ws['E1'] = 'Corresponding Price'

row = 2
for crypto in cryptos[:5]:
    name = crypto.find('p', {'class': 'sc-1eb5slv-0'}).text
    symbol = crypto.find('p', {'class': 'sc-1eb5slv-1'}).text
    price = crypto.find('div', {'class': 'price___3rj7O'}).text
    percent_change = crypto.find('div', {'class': 'sc-1yv6u5n-0'}).text
    corresponding_price = 'N/A'  # Calculate based on % change
    ws.cell(row=row, column=1, value=name)
    ws.cell(row=row, column=2, value=symbol)
    ws.cell(row=row, column=3, value=price)
    ws.cell(row=row, column=4, value=percent_change)
    ws.cell(row=row, column=5, value=corresponding_price)
    row += 1

# Apply formatting to the spreadsheet
for col in ['A', 'B', 'C', 'D', 'E']:
    ws.column_dimensions[col].width = 20
    ws[f'{col}1'].font = openpyxl.styles.Font(bold=True)
# ws['C:C'].number_format = '$0.00'
# ws['D:D'].number_format = '0.00%'

# Check Bitcoin and Ethereum value every 5 minutes and send text message if it changes by $5
account_sid = 'YOUR_ACCOUNT_SID'
auth_token = 'YOUR_AUTH_TOKEN'
client = Client(account_sid, auth_token)


def check_price():
    btc_price = float(ws.cell(row=2, column=3).value.replace(
        '$', '').replace(',', ''))
    eth_price = float(ws.cell(row=3, column=3).value.replace(
        '$', '').replace(',', ''))
    btc_prev_price = btc_price
    eth_prev_price = eth_price
    while True:
        response = requests.get(url)
        soup = BeautifulSoup(response.content, 'html')
